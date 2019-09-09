import httplib2
import os
import sys
import openpyxl
from openpyxl import load_workbook

from apiclient.discovery import build
from oauth2client.client import flow_from_clientsecrets
from oauth2client.file import Storage
from oauth2client.tools import argparser, run_flow

# Remixed from: https://developers.google.com/youtube/v3/code_samples/python#retrieve_my_uploads
# Some of the original code's comments are still included

# The CLIENT_SECRETS_FILE variable specifies the name of a file that contains
# the OAuth 2.0 information for this application, including its client_id and
# client_secret. You can acquire an OAuth 2.0 client ID and client secret from
# the {{ Google Cloud Console }} at
# {{ https://cloud.google.com/console }}.
# Please ensure that you have enabled the YouTube Data API for your project.
# For more information about using OAuth2 to access the YouTube Data API, see:
#   https://developers.google.com/youtube/v3/guides/authentication
# For more information about the client_secrets.json file format, see:
#   https://developers.google.com/api-client-library/python/guide/aaa_client_secrets
CLIENT_SECRETS_FILE = "client_secret.json"

# This variable defines a message to display if the CLIENT_SECRETS_FILE is
# missing.
MISSING_CLIENT_SECRETS_MESSAGE = """
WARNING: Please configure OAuth 2.0

To make this sample run you will need to populate the client_secrets.json file
found at:

   %s

with information from the {{ Cloud Console }}
{{ https://cloud.google.com/console }}

For more information about the client_secrets.json file format, please visit:
https://developers.google.com/api-client-library/python/guide/aaa_client_secrets
""" % os.path.abspath(os.path.join(os.path.dirname(__file__),
                                   CLIENT_SECRETS_FILE))

# This OAuth 2.0 access scope allows for read-only access to the authenticated
# user's account, but not other types of account access.
YOUTUBE_READONLY_SCOPE = "https://www.googleapis.com/auth/youtube.readonly"
YOUTUBE_API_SERVICE_NAME = "youtube"
YOUTUBE_API_VERSION = "v3"

flow = flow_from_clientsecrets(CLIENT_SECRETS_FILE,
  message=MISSING_CLIENT_SECRETS_MESSAGE,
  scope=YOUTUBE_READONLY_SCOPE)

storage = Storage("%s-oauth2.json" % sys.argv[0])
credentials = storage.get()

if credentials is None or credentials.invalid:
  flags = argparser.parse_args()
  credentials = run_flow(flow, storage, flags)

youtube = build(YOUTUBE_API_SERVICE_NAME, YOUTUBE_API_VERSION,
  http=credentials.authorize(httplib2.Http()))

"""
FYI for the Eurovision account:
This channel's ID is UCRpjHHu8ivVWs73uxHlWwFA. Its title is Eurovision Song Contest, and it has 2453330071 views. Its uploads playlist ID is UURpjHHu8ivVWs73uxHlWwFA.
"""

# Retrieve the contentDetails part of the channel resource for the
# authenticated user's channel.
channels_response = youtube.channels().list(
  part="contentDetails",
  forUsername="Eurovision"
).execute()

uploads_list_id = channels_response["items"][0]["contentDetails"]["relatedPlaylists"]["uploads"]

print("Videos in list %s" % uploads_list_id)

# Retrieve the list of videos uploaded to the authenticated user's channel.
playlistitems_list_request = youtube.playlistItems().list(
  playlistId=uploads_list_id,
  part="snippet",
  maxResults=50
)

def xlref(row, column): #gets row and column indices and converts to Excel cell name
    #Code adapted from https://stackoverflow.com/questions/31420817/convert-excel-row-column-indices-to-alphanumeric-cell-reference-in-python-openpy
    return openpyxl.utils.get_column_letter(column) + str(row)

wb = openpyxl.Workbook()
dest_filename = 'eurovision-youtube-videos.xlsx'
ws1 = wb.active
ws1.title = 'esc youtube vids'
header = ['ID', 'TITLE', 'URL', 'UPLOADED', 'VIEWS', 'LIKES', 'DISLIKES']
ws1.append(header)
total_so_far = 0
wb.save(dest_filename)

while playlistitems_list_request:
  wb = load_workbook(dest_filename)
  ws1 = wb.active

  playlistitems_list_response = playlistitems_list_request.execute()

  # Print information about each video.
  for playlist_item in playlistitems_list_response["items"]:
    entry = []
    title = playlist_item["snippet"]["title"]
    print("Processing",title)
    video_id = playlist_item["snippet"]["resourceId"]["videoId"]
    video_url = "http://youtube.com/watch?v="+str(video_id)
    entry.append(video_id)
    entry.append(title)
    entry.append(video_url)
    upload_time = playlist_item["snippet"]["publishedAt"]
    entry.append(upload_time)
    videoitem_request = youtube.videos().list(
      part="statistics,snippet",
      id=video_id
    )
    try:
      videoitem = videoitem_request.execute()
      view_count = videoitem["items"][0]["statistics"]["viewCount"]
      entry.append(view_count)
      like_count = videoitem["items"][0]["statistics"]["likeCount"]
      entry.append(like_count)
      dislike_count = videoitem["items"][0]["statistics"]["dislikeCount"]
      entry.append(dislike_count)
    except KeyError:
      print("I guess something's wrong with", title)
      for oops in range(0,3):
        entry.append(' ')
    ws1.append(entry)
    total_so_far += 1

  playlistitems_list_request = youtube.playlistItems().list_next(
    playlistitems_list_request, playlistitems_list_response)

  wb.save(dest_filename)
  print(total_so_far-1,"videos saved")