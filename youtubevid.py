#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Housekeeping: do not edit
import json, io, tweepy, time, urllib3
from random import randint

import openpyxl
import random

def xlref(row, column): #gets row and column indices and converts to Excel cell name
    #Code adapted from https://stackoverflow.com/questions/31420817/convert-excel-row-column-indices-to-alphanumeric-cell-reference-in-python-openpy
    return openpyxl.utils.get_column_letter(column) + str(row)

def youtubetweet():
	# Generates a tweet with a YouTube video from the Eurovision channel with its title and the Eurovision hashtag

	wb = openpyxl.load_workbook('esc-youtube.xlsx') #opens our Excel Workbook
	sheet = wb[('esc youtube vids')] #retrieves the specific Sheet we want

	# Picks a random number
	rand_row = randint(1, sheet.max_row+1 - 1)
    
	# Gets the title and URL of a video based on the random number
	title_cell = xlref(rand_row,2)
	url_cell = xlref(rand_row,3)

	# Creates the tweet and posts it
	title = sheet[title_cell].value
	url = sheet[url_cell].value
	my_tweet = (f"{title} #Eurovision {url}")
	return my_tweet