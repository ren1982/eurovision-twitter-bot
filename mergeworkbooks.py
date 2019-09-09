import httplib2
import os
import sys
import openpyxl
from openpyxl import load_workbook
from itertools import islice

wb1 = load_workbook('eurovision-youtube-videos_old.xlsx') # old version of the worksheet
ws1 = wb1[('esc youtube vids')]
mostrecent = ws1['A2'].value # gets the video id of the most recent video found in the old worksheet
wb2 = load_workbook('eurovision-youtube-videos.xlsx') # latest version of the worksheet
ws2 = wb2[('esc youtube vids')]
for i in range(1,ws2.max_row): # find where the most recent video in the old worksheet is located in the new worksheet to find where copying the value ends
	cell = 'A' + str(i)
	curr_id = ws2[cell].value
	if mostrecent == curr_id:
		end = i-1
		break
for row in islice(ws2.rows,2,end):
	ws1.append((cell.value for cell in row))
wb1.save('esc-youtube.xlsx')